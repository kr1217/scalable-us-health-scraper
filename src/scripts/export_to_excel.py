import pandas as pd
import asyncio
from sqlalchemy import select
from datetime import datetime
import os
import sys

# Add src to path
sys.path.append(os.path.abspath("."))

from src.core.database import AsyncSessionLocal
from src.core.orm_models import LeadORM

async def export_leads_to_excel(file_path: str = "health_leads_formatted.xlsx"):
    """
    Fetches all leads from PostgreSQL and exports/appends them to the Excel file.
    """
    print(f"Exporting leads to {file_path}...")
    
    async with AsyncSessionLocal() as session:
        stmt = select(LeadORM).order_by(LeadORM.extracted_at.desc())
        result = await session.execute(stmt)
        leads = result.scalars().all()
        
        if not leads:
            print("No leads found in database to export.")
            return

        # Prepare list of dictionaries for Pandas
        data = []
        for lead in leads:
            data.append({
                "First Name": lead.first_name,
                "Last Name": lead.last_name,
                "Date of Birth": lead.date_of_birth,
                "Phone Number": lead.phone_number,
                "Phone Source": lead.phone_number_source,
                "State": lead.state,
                "City": lead.city,
                "Address": lead.address,
                "Disease History": lead.disease_history,
                "Source": lead.source,
                "Source URL": lead.source_url,
                "Extracted At": lead.extracted_at.strftime("%Y-%m-%d %H:%M:%S") if lead.extracted_at else None
            })
            
        df_new = pd.DataFrame(data)
        
        columns = [
            'First Name', 'Last Name', 'Date of Birth', 'Phone Number', 
            'Phone Source', 'State', 'City', 'Address', 'Disease History', 
            'Source', 'Source URL', 'Extracted At'
        ]
        
        try:
            import openpyxl
            if os.path.exists(file_path):
                # Load existing workbook to preserve custom formatting and column widths
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Capture the user's custom headers from row 1
                header = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
                
                # Safely delete old data rows while preserving column layout rules
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
                    
                # Insert the new leads into the correct columns
                for row_idx, lead in enumerate(data, start=2):
                    for col_name in columns:
                        if col_name in header:
                            col_idx = header.index(col_name) + 1
                            ws.cell(row=row_idx, column=col_idx, value=lead.get(col_name))
                
                wb.save(file_path)
                print(f"Successfully synced {len(data)} leads into {file_path} (Formatting Preserved)")
            else:
                # Fallback to Pandas if the file doesn't exist yet
                df_new = pd.DataFrame(data)[columns]
                df_new.to_excel(file_path, index=False)
                print(f"Successfully generated {len(data)} leads to {file_path}")
                
        except Exception as e:
            print(f"Error saving to Excel: {e}")

if __name__ == "__main__":
    asyncio.run(export_leads_to_excel())
